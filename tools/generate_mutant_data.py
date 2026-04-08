#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成变异生物和野外Boss数据表Excel文件
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# 地图配置
MAPS = {
    1: {'name': '辐射废墟', 'element': 'dark', 'mutantElement': '暗'},
    2: {'name': '有毒沼泽', 'element': 'water', 'mutantElement': '水'},
    3: {'name': '钢铁要塞', 'element': 'dark', 'mutantElement': '暗'},
    4: {'name': '冰封高原', 'element': 'water', 'mutantElement': '水'},
    5: {'name': '炽热火山', 'element': 'fire', 'mutantElement': '火'},
    6: {'name': '幽暗森林', 'element': 'wind', 'mutantElement': '风'},
    7: {'name': '沙漠荒原', 'element': 'fire', 'mutantElement': '火'},
    8: {'name': '深海沉城', 'element': 'water', 'mutantElement': '水'},
    9: {'name': '机械迷宫', 'element': 'dark', 'mutantElement': '暗'},
    10: {'name': '虚空裂隙', 'element': 'dark', 'mutantElement': '暗'},
}

ELEMENTS = ['dark', 'water', 'fire', 'wind', 'light']
PROFESSIONS = ['tank', 'dps', 'support']

# 10个野外Boss数据
WILD_BOSSES = [
    {
        'bossId': 'WILDBOSS_001',
        'name': '辐射巨兽·荒原咆哮者',
        'mapId': 1,
        'element': 'dark',
        'baseHp': 8000,
        'baseAtk': 450,
        'baseSpd': 35,
        'description': '核战后的废墟中咆哮的巨型变异生物',
        'portrait': 'radiation_beast'
    },
    {
        'bossId': 'WILDBOSS_002',
        'name': '剧毒领主·沼泽窒息者',
        'mapId': 2,
        'element': 'water',
        'baseHp': 7500,
        'baseAtk': 480,
        'baseSpd': 38,
        'description': '沼泽深处的剧毒统治者',
        'portrait': 'toxic_lord'
    },
    {
        'bossId': 'WILDBOSS_003',
        'name': '钢铁巨像·锈蚀泰坦',
        'mapId': 3,
        'element': 'dark',
        'baseHp': 9000,
        'baseAtk': 420,
        'baseSpd': 25,
        'description': '废弃军事要塞中的古老战争机甲',
        'portrait': 'steel_colossus'
    },
    {
        'bossId': 'WILDBOSS_004',
        'name': '冰霜巨龙·永冻脊骨',
        'mapId': 4,
        'element': 'water',
        'baseHp': 8500,
        'baseAtk': 500,
        'baseSpd': 40,
        'description': '冰封高原的上古巨龙遗骸',
        'portrait': 'frost_dragon'
    },
    {
        'bossId': 'WILDBOSS_005',
        'name': '熔岩古龙·灰烬焚尽者',
        'mapId': 5,
        'element': 'fire',
        'baseHp': 8800,
        'baseAtk': 550,
        'baseSpd': 36,
        'description': '炽热火山深处的火焰精灵化身',
        'portrait': 'lava_dragon'
    },
    {
        'bossId': 'WILDBOSS_006',
        'name': '精灵古树·暗夜吞噬者',
        'mapId': 6,
        'element': 'wind',
        'baseHp': 9500,
        'baseAtk': 400,
        'baseSpd': 20,
        'description': '幽暗森林最古老的巨树意志',
        'portrait': 'ancient_tree'
    },
    {
        'bossId': 'WILDBOSS_007',
        'name': '沙海蜃楼·蜃景幻主',
        'mapId': 7,
        'element': 'fire',
        'baseHp': 7000,
        'baseAtk': 520,
        'baseSpd': 45,
        'description': '沙漠荒原中的神秘蜃景存在',
        'portrait': 'mirage_lord'
    },
    {
        'bossId': 'WILDBOSS_008',
        'name': '海皇利维坦·深渊沉眠者',
        'mapId': 8,
        'element': 'water',
        'baseHp': 9200,
        'baseAtk': 480,
        'baseSpd': 32,
        'description': '沉没都市的真正统治者',
        'portrait': 'sea_leviathan'
    },
    {
        'bossId': 'WILDBOSS_009',
        'name': '超级计算机·矩阵母体',
        'mapId': 9,
        'element': 'dark',
        'baseHp': 10000,
        'baseAtk': 380,
        'baseSpd': 15,
        'description': '觉醒的超级AI，将入侵者视为病毒',
        'portrait': 'matrix_mother'
    },
    {
        'bossId': 'WILDBOSS_010',
        'name': '维度撕裂者·虚空信使',
        'mapId': 10,
        'element': 'dark',
        'baseHp': 8500,
        'baseAtk': 560,
        'baseSpd': 50,
        'description': '来自异次元的侦察信使',
        'portrait': 'void_herald'
    },
]

# 最终Boss
FINAL_BOSS = {
    'bossId': 'BOSS_FINAL_001',
    'name': '终焉之核·奥米茄',
    'mapId': 10,
    'zoneId': 'FINAL_ZONE',
    'element': 'dark',
    'baseHp': 15000,
    'baseAtk': 600,
    'baseSpd': 40,
    'phases': 3,
    'description': '融合技术的创造者，唯一成功的完美融合体',
    'portrait': 'omega_core'
}

# 生成变异生物数据
def generate_mutants():
    mutants = []
    idx = 1
    
    for map_id in range(1, 11):
        map_name = MAPS[map_id]['name']
        mutant_element = MAPS[map_id]['mutantElement']
        
        # 为每个地图生成5元素×3职业=15种变异生物
        for elem_idx, element in enumerate(ELEMENTS):
            for prof_idx, profession in enumerate(PROFESSIONS):
                # 基础属性根据职业和元素调整
                if profession == 'tank':
                    base_hp = 1200 + map_id * 100
                    base_atk = 80 + map_id * 10
                    base_spd = 25 + map_id * 2
                elif profession == 'dps':
                    base_hp = 800 + map_id * 80
                    base_atk = 120 + map_id * 15
                    base_spd = 35 + map_id * 3
                else:  # support
                    base_hp = 900 + map_id * 90
                    base_atk = 100 + map_id * 12
                    base_spd = 30 + map_id * 2
                
                # 元素前缀名称
                elem_prefix = {
                    'dark': ['暗影', '虚空', '腐蚀', '毒素', '死灵'],
                    'water': ['深渊', '潮汐', '海流', '寒冰', '暴雨'],
                    'fire': ['烈焰', '熔岩', '灰烬', '火山', '燃烧'],
                    'wind': ['风暴', '疾风', '雷霆', '飓风', '龙卷'],
                    'light': ['圣光', '雷霆', '光辉', '烈日', '星尘']
                }
                
                # 职业名称
                prof_names = {
                    'tank': ['巨兽', '甲虫', '甲壳', '巨怪', '泰坦'],
                    'dps': ['猎手', '猛兽', '利爪', '尖牙', '利刃'],
                    'support': ['精灵', '守护', '使者', '精魂', '灵体']
                }
                
                name_idx = (elem_idx * 3 + prof_idx) % 5
                elem_name = elem_prefix[element][name_idx]
                prof_name = prof_names[profession][name_idx]
                
                mutant_id = f"MUT_{map_id:02d}_{element[:2].upper()}_{profession[:2].upper()}_{idx:02d}"
                
                mutants.append({
                    'enemyId': mutant_id,
                    'name': f'{elem_name}{prof_name}',
                    'mapId': map_id,
                    'element': element,
                    'profession': profession,
                    'type': 'mutant',
                    'baseHp': base_hp,
                    'baseAtk': base_atk,
                    'baseSpd': base_spd,
                    'description': f'{map_name}的{elem_name}{prof_name}',
                    'portrait': f'mutant_{map_id:02d}_{element[:2]}_{profession[:2]}_{idx:02d}'
                })
                idx += 1
    
    return mutants

def create_excel_files():
    output_dir = 'd:/77-myProject/苏打地牢/assets/data/excel'
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. 创建野外Boss数据表
    wb_bosses = Workbook()
    ws_boss = wb_bosses.active
    ws_boss.title = '野外Boss'
    
    # 添加表头
    headers = ['bossId', 'name', 'mapId', 'element', 'baseHp', 'baseAtk', 'baseSpd', 'description', 'portrait']
    for col, header in enumerate(headers, 1):
        ws_boss.cell(1, col, header)
        ws_boss.cell(2, col, 'string')
        ws_boss.cell(3, col, '野外Boss唯一ID')
    
    # 添加数据
    for row, boss in enumerate(WILD_BOSSES, 4):
        ws_boss.cell(row, 1, boss['bossId'])
        ws_boss.cell(row, 2, boss['name'])
        ws_boss.cell(row, 3, boss['mapId'])
        ws_boss.cell(row, 4, boss['element'])
        ws_boss.cell(row, 5, boss['baseHp'])
        ws_boss.cell(row, 6, boss['baseAtk'])
        ws_boss.cell(row, 7, boss['baseSpd'])
        ws_boss.cell(row, 8, boss['description'])
        ws_boss.cell(row, 9, boss['portrait'])
    
    # 添加最终Boss
    final_row = len(WILD_BOSSES) + 4
    ws_boss.cell(final_row, 1, FINAL_BOSS['bossId'])
    ws_boss.cell(final_row, 2, FINAL_BOSS['name'])
    ws_boss.cell(final_row, 3, FINAL_BOSS['mapId'])
    ws_boss.cell(final_row, 4, FINAL_BOSS['element'])
    ws_boss.cell(final_row, 5, FINAL_BOSS['baseHp'])
    ws_boss.cell(final_row, 6, FINAL_BOSS['baseAtk'])
    ws_boss.cell(final_row, 7, FINAL_BOSS['baseSpd'])
    ws_boss.cell(final_row, 8, FINAL_BOSS['description'])
    ws_boss.cell(final_row, 9, FINAL_BOSS['portrait'])
    
    wb_bosses.save(f'{output_dir}/wildBosses.xlsx')
    print(f'✓ 已创建野外Boss数据表: wildBosses.xlsx (包含{len(WILD_BOSSES) + 1}个Boss)')
    
    # 2. 创建变异生物数据表
    mutants = generate_mutants()
    
    wb_mutants = Workbook()
    ws_mut = wb_mutants.active
    ws_mut.title = '变异生物'
    
    # 添加表头
    headers = ['enemyId', 'name', 'mapId', 'element', 'profession', 'type', 'baseHp', 'baseAtk', 'baseSpd', 'description', 'portrait']
    for col, header in enumerate(headers, 1):
        ws_mut.cell(1, col, header)
        ws_mut.cell(2, col, 'string')
        ws_mut.cell(3, col, f'变异生物{header}')
    
    # 添加数据
    for row, mutant in enumerate(mutants, 4):
        ws_mut.cell(row, 1, mutant['enemyId'])
        ws_mut.cell(row, 2, mutant['name'])
        ws_mut.cell(row, 3, mutant['mapId'])
        ws_mut.cell(row, 4, mutant['element'])
        ws_mut.cell(row, 5, mutant['profession'])
        ws_mut.cell(row, 6, mutant['type'])
        ws_mut.cell(row, 7, mutant['baseHp'])
        ws_mut.cell(row, 8, mutant['baseAtk'])
        ws_mut.cell(row, 9, mutant['baseSpd'])
        ws_mut.cell(row, 10, mutant['description'])
        ws_mut.cell(row, 11, mutant['portrait'])
    
    wb_mutants.save(f'{output_dir}/mutants.xlsx')
    print(f'✓ 已创建变异生物数据表: mutants.xlsx (包含{len(mutants)}个变异生物)')
    
    # 3. 生成怪物名称映射表（用于生成脚本）
    mapping = {}
    
    # 添加野外Boss
    for boss in WILD_BOSSES:
        mapping[boss['portrait']] = {
            'type': 'wildBoss',
            'name': boss['name'],
            'prompt': f"【Boss名称】{boss['name']}\n【描述】{boss['description']}"
        }
    
    # 添加最终Boss
    mapping[FINAL_BOSS['portrait']] = {
        'type': 'finalBoss',
        'name': FINAL_BOSS['name'],
        'prompt': f"【Boss名称】{FINAL_BOSS['name']}\n【描述】{FINAL_BOSS['description']}"
    }
    
    # 添加变异生物
    for mutant in mutants:
        mapping[mutant['portrait']] = {
            'type': 'mutant',
            'name': mutant['name'],
            'mapId': mutant['mapId'],
            'element': mutant['element'],
            'profession': mutant['profession'],
            'prompt': f"【变异生物名称】{mutant['name']}\n【地图】{MAPS[mutant['mapId']]['name']}\n【元素】{mutant['element']}\n【职业】{mutant['profession']}"
        }
    
    # 保存映射表为JSON
    import json
    with open('d:/77-myProject/苏打地牢/tools/mutant_portrait_mapping.json', 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)
    
    print(f'✓ 已创建怪物立绘映射表: mutant_portrait_mapping.json (包含{len(mapping)}个映射)')
    print(f'\n📊 总计:')
    print(f'  - 野外Boss: {len(WILD_BOSSES)}个')
    print(f'  - 最终Boss: 1个')
    print(f'  - 变异生物: {len(mutants)}个')
    print(f'  - 合计立绘: {len(mapping)}个')

if __name__ == '__main__':
    create_excel_files()
